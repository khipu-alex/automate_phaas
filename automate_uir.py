import pandas as pd
import matplotlib.pyplot as plt
import glob
import numpy as np
from matplotlib.colors import ListedColormap
import seaborn as sns
from matplotlib.ticker import FuncFormatter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def main():
    excel_files = [file for file in glob.glob('*.xlsx') if file != "phish_automation_results.xlsx"]
    excel_file = excel_files[0]
    xls = pd.ExcelFile(excel_file)
    first_sheet = xls.sheet_names[0]
    df = pd.read_excel(excel_file, sheet_name=first_sheet)
    col_names = [col.lower() for col in df.columns.tolist()]
    useful_cols = ["staff_type", "department", "dept", "employee_subgroup", "subgroup", "location", "division", "team"]

    with pd.ExcelWriter("phish_automation_results.xlsx", engine='openpyxl') as writer:
        write_to_new_sheet(writer)
        for name in useful_cols:
            if name in col_names:
                col_data(writer, excel_file, name)

        multi_campaign_info(writer)
        repeat_offenders_data(writer)
        pass_fail(writer)
        os_browser(writer)

        for sheet_name in writer.sheets:
            sheet = writer.sheets[sheet_name]
            style_sheet(sheet)

def style_sheet(sheet):
    header_font = Font(bold=True, color="FFFFFF", size=16)
    header_fill = PatternFill(start_color="3b3e3f", end_color="3b3e3f", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    data_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_style

    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass

    adjusted_width = max_length + 5
    sheet.column_dimensions[column].width = adjusted_width

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = border_style
            cell.fill = data_fill

def write_to_new_sheet(writer):
    excel_files = [file for file in glob.glob('*.xlsx') if file != "phish_automation_results.xlsx"]
    excel_file = excel_files[0]
    xls = pd.ExcelFile(excel_file)

    print(f"\nPROCESSING RAW DATA")

    columns_to_remove = ['phone', 'gender', 'lure_submitted_at', 'mail_submitted_at', 'reported', 'reported_at', 'domain', 'scenario', 
                        'comment', 'proxy_ip', 'success_rate', 'click_rate', 'scenario_time', 'awareness_time',
                        'awareness_clicked_at', 'email_subject', 'first_click_after_delivery', 'first_report_after_delivery', 
                        'out_of_office_at', 'bounced_at', 'responded_at', 'certificate_received', 'training_with_quiz_passed', 
                        'training_with_quiz_passed_at', 'training_noquiz_finished', 'training_noquiz_finished_at', 'recipient_branches', 
                        'reminder_click_submitted_at', 'reminder_training_start_submitted_at', 'reminder_training_finish_submitted_at', 
                        'answers_count_0', 'answers_percent_0', 'wrong_answers_count_0', 'wrong_answers_percent_0', 
                        'training_succeeded_0', 'training_succeeded_at_0']  

    new_columns = ['name', 'email', 'link', 'employee_subgroup', 'team', 'dept', 'clicked', 'clicked_at', 'succeeded', 'succeeded_at', 
                   'trained', 'trained_at', 'staff_type', 'location', 'division', 'os', 'ip', 'browser', 'plugins', 'country', 'downloaded_files', 
                   'collected_data', 'correct_answers_count_0', 'correct_answers_percent_0', 'quiz_time_spent_0']
    
    new_column_names = ['Name', 'Email', 'Link', 'Employee Subgroup', 'Team', 'Dept', 'Clicked', 'Clicked At', 'Succeeded', 'Succeeded At',
                       'Trained', 'Trained At', 'Staff Type', 'Location', 'Division', 'OS', 'IP', 'Browser', 'Plugins', 'Country', 'Downloaded Files', 
                       'Collected Data', 'Correct Answers Count', 'Correct Answers Percent', 'Quiz Time Spent']  

    sheet = xls.sheet_names[0]
    df = pd.read_excel(excel_file, sheet_name=sheet)
    df = df.drop(columns=columns_to_remove, errors='ignore')  

    df['link'] = "https://[DOMAIN].khipuawareness.com/awareness/v/" + df['link'] + "/index.html"

    for i, col in enumerate(new_columns):
        if col not in df.columns:
            df.insert(i, col, pd.NA)

    if len(df.columns) == len(new_column_names):
        df.columns = new_column_names
    else:
        print(f"Warning: Column count mismatch for sheet '{sheet}'")

    df = df.dropna(axis=1, how='all')
    df = df.loc[:, df.ne('-').any()]

    df.to_excel(writer, sheet_name='Raw Data', index=False)

def multi_campaign_info(writer):
    try:
        excel_files = [file for file in glob.glob('*.xlsx') if file != "phish_automation_results.xlsx"]
        excel_file = excel_files[0]

        print(f"\nPROCESSING: {excel_file}, for: Multi-Campaign Stats")

        total_clicks = 0
        total_clicks_percent = 0
        total_submit = 0
        total_submit_percent = 0
        total_employees = 0

        # Open the Excel file
        xls = pd.ExcelFile(excel_file)
        for sheet in xls.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet)

            # Ensure there's an 'Email' column
            total_col = df.get('email')
            if total_col is None:
                continue

            # Count all non-null 'Email' entries for Total Employees
            total_employees += total_col.notna().sum()

            # Count 'Clicked == "Y"'
            if 'clicked' in df.columns:
                total_clicks += (df['clicked'] == "Y").sum()

            # Count 'Succeeded == "Y"'
            if 'succeeded' in df.columns:
                total_submit += (df['succeeded'] == "Y").sum()

        total_clicks_percent = (total_clicks / total_employees) * 100
        total_submit_percent = (total_submit / total_employees) * 100 

        # Create combined DataFrame for output
        combined_df = pd.DataFrame({
            'Total Employees': [total_employees],
            'Total Clicked': [total_clicks],
            'Total Submitted': [total_submit],
            'Total Clicked Percentage': [total_clicks_percent],
            'Total Submit Percentage': [total_submit_percent]
        })

        combined_df.to_excel(writer, sheet_name='Multi Campaign Stats', index=False)

    except Exception as e:
        print(f"\nError: {e}\n")

def repeat_offenders_data(writer):
    try:
        excel_files = [file for file in glob.glob('*.xlsx') if file != "phish_automation_results.xlsx"]

        if len(excel_files) > 1:
                
            print("\nThis file only takes in two excel files, if you have more than two campaigns, it will ONLY get repeat offenders across ALL excel files, ensure this is correct.")

            email_lists_clicked = []

            for file in excel_files:
                print(f"Processing {file}")
                df = pd.read_excel(file)

                if 'email' in df.columns and 'clicked' in df.columns and 'succeeded' in df.columns:
                    df['source_file'] = file
                    clicked_df = df[df['clicked'] == 'Y'][['email', 'clicked', 'succeeded', 'source_file']]
                    email_lists_clicked.append(clicked_df)

            
            all_emails = pd.concat(email_lists_clicked, ignore_index=True)
            duplicate_emails = all_emails[all_emails.duplicated(subset=['email'], keep=False)]
            pivoted = duplicate_emails.pivot_table(index='email', columns='source_file', values=['clicked', 'succeeded'], aggfunc=lambda x: ' | '.join(x))
            pivoted.columns = [f'{col[0]}_{col[1]}' for col in pivoted.columns]
            pivoted = pivoted.reset_index()

            pivoted.to_excel(writer, sheet_name='Repeat Offenders', index=False)

        else:
            print("\nSkipping repeat offenders, only 1 suitable file was found in the directory")

    except Exception as e:
        print(f"\nError: {e}\n")

def pass_fail(writer):
    try:
        excel_files = [file for file in glob.glob('*.xlsx') if file != "phish_automation_results.xlsx"]
        excel_file = excel_files[0]

        print(f"\nPROCESSING: {excel_file}, FOR: PASS/FAIL DATA")

        passes = []  

        xls = pd.ExcelFile(excel_file)
        for sheet in xls.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet)
            if 'trained' in df.columns:
                trained_df = df[df['trained'] == "Y"]

                if 'answers_percent_0' in df.columns:
                    percent_col = trained_df['answers_percent_0'].dropna().tolist()

            if 'succeeded' in df.columns:
                succeeded_count = df[df['succeeded'] == "Y"]
                

            passes.extend(percent_col)

        df = pd.DataFrame(percent_col)

        total_count = df.shape[0]
        count_of_awareness_opened = (total_count / succeeded_count.shape[0]) * 100
        over_75_count = (df[df[0] > 0.7].shape[0])
        average = df[0].mean() * 100
        not_passed_count = total_count - over_75_count
        
        awareness = pd.DataFrame({
            'Total Awareness emails sent': [succeeded_count.shape[0]],
            'Awareness emails sent and opened': [total_count],
            'Awareness sent and opened %': count_of_awareness_opened,
            'Not Passed': [not_passed_count],
            'Passed': [over_75_count],
            'Average Score': [average]
        })

        
        awareness.to_excel(writer, sheet_name='Awareness Results', index=False)

    except Exception as e:
        print(f"\nError: {e}\n")

def os_browser(writer):
    try:
        excel_files = [file for file in glob.glob('*.xlsx') if file != "phish_automation_results.xlsx"]
        excel_file = excel_files[0]

        print(f"\nPROCESSING: {excel_file}, FOR: OS/BROWSER DATA")

        os = []
        browser = []

        xls = pd.ExcelFile(excel_file)

        for sheet in xls.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet)
            
            if 'succeeded' in df.columns:
                df = df[df['succeeded'] == "Y"]
                if 'os' in df.columns and 'browser' in df.columns:
                    os.extend(df['os'].tolist())
                    browser.extend(df['browser'].tolist())

        browser_counts = pd.Series(browser).value_counts()
        os_counts = pd.Series(os).value_counts()

        os_df = pd.DataFrame({
            'OS': os_counts.index,
            'OS Count': os_counts.values
        })

        browser_df = pd.DataFrame({
            'Browser': browser_counts.index,
            'Browser Count': browser_counts.values,
        })

        os_df.to_excel(writer, sheet_name='OS Data', index=False) 
        browser_df.to_excel(writer, sheet_name='Browser Data', index=False) 
    
    except Exception as e:
        print(f"Error: {e}")
        
def col_data(writer, excel_file, col):
    try:
        print(f"\nPROCESSING: {excel_file}, FOR: {col}")

        click = []
        submit = []
        all_record = []

        # Load the Excel file
        xls = pd.ExcelFile(excel_file)
        
        for sheet in xls.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet)
            
            if col in df.columns:
                col_data = df[col].dropna()

                all_record.extend(col_data)

                if 'clicked' in df.columns:
                    click.extend(col_data[df['clicked'] == "Y"])

                if 'succeeded' in df.columns:
                    submit.extend(col_data[df['succeeded'] == "Y"])

        combined_df = pd.DataFrame({
            'Clicked': pd.Series(click).value_counts(),
            'Submitted': pd.Series(submit).value_counts(),
            'Total Employees': pd.Series(all_record).value_counts()
        }).fillna(0).astype(int)

        combined_df['Total'] = combined_df.index.astype(str) + ' (' + combined_df['Total Employees'].astype(str) + ')'
        combined_df['Clicked Percentage'] = (combined_df['Clicked'] / combined_df['Total Employees'].replace(0, np.nan)) * 100
        combined_df['Submitted Percentage'] = (combined_df['Submitted'] / combined_df['Total Employees'].replace(0, np.nan)) * 100

        combined_df.reset_index(drop=False, inplace=True)
        combined_df.rename(columns={'index': f'{col[0].upper()}{col[1:]}'}, inplace=True)
        filtered_combined_df = combined_df[combined_df['Total Employees'] > 10]

        if filtered_combined_df.iloc[0,0] == "-":
            print(f"Skipping {col}")
        else:
            new_name = ""
            col_parts = col.split("_")
            new_name = " ".join([item.capitalize() for item in col_parts])

            filtered_combined_df.to_excel(writer, sheet_name=new_name, index=False) 

    except Exception as e:
        print(f"\nError: {e}\n")



if __name__ == "__main__":
    main()